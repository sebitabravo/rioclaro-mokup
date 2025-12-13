from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db.models import Q, Avg, Min, Max, Count
from django.db import transaction
from rest_framework import generics, status, filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import rest_framework as filters_rf
from django.http import HttpResponse
import datetime
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import (
    Measurement,
    Threshold,
    Alert,
    MeasurementConfiguration,
    MeasurementType,
    AlertStatus
)
from .serializers import (
    MeasurementListSerializer,
    MeasurementDetailSerializer,
    MeasurementCreateSerializer,
    LatestMeasurementSerializer,
    ThresholdSerializer,
    AlertSerializer,
    AlertActionSerializer,
    MeasurementConfigurationSerializer,
    MeasurementStatsSerializer,
    BatchMeasurementCreateSerializer,
    # Módulo 3: Reportes - Serializers
    DailyAverageReportSerializer,
    CriticalEventsReportSerializer,
    ComparativeReportSerializer
)
from stations.models import Station
from users.models import UserRole


class MeasurementPagination(PageNumberPagination):
    """Paginación personalizada para mediciones"""
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 1000


class MeasurementFilter(filters_rf.FilterSet):
    """Filtros para mediciones (RF2.3)"""
    date_from = filters_rf.DateTimeFilter(field_name='timestamp', lookup_expr='gte')
    date_to = filters_rf.DateTimeFilter(field_name='timestamp', lookup_expr='lte')
    date_range = filters_rf.DateFromToRangeFilter(field_name='timestamp')
    measurement_type = filters_rf.MultipleChoiceFilter(choices=MeasurementType.choices)
    station = filters_rf.NumberFilter(field_name='station__id')
    sensor = filters_rf.NumberFilter(field_name='sensor__id')
    quality_flag = filters_rf.ChoiceFilter(choices=[
        ('good', 'Buena'),
        ('suspect', 'Sospechosa'),
        ('poor', 'Mala'),
        ('missing', 'Faltante'),
    ])

    class Meta:
        model = Measurement
        fields = ['date_from', 'date_to', 'measurement_type', 'station', 'sensor', 'quality_flag']


# RF2.1 - Datos en Tiempo Real
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def latest_measurement_by_station(request, station_id):
    """
    RF2.1: Obtiene la última medición de una estación específica

    Endpoint: GET /api/measurements/stations/{station_id}/latest/
    """
    try:
        station = get_object_or_404(Station, id=station_id, is_active=True)

        # Verificar permisos: usuarios solo pueden ver estaciones asignadas (excepto admins)
        if request.user.role != UserRole.ADMIN:
            if not station.assigned_users.filter(id=request.user.id).exists():
                return Response(
                    {'error': 'No tiene permisos para acceder a esta estación'},
                    status=status.HTTP_403_FORBIDDEN
                )

        # Obtener la última medición
        latest_measurement = Measurement.objects.filter(
            station=station
        ).select_related('station', 'sensor').order_by('-timestamp').first()

        if not latest_measurement:
            return Response(
                {
                    'station_id': station_id,
                    'station_name': station.name,
                    'message': 'No hay mediciones disponibles para esta estación'
                },
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = LatestMeasurementSerializer(latest_measurement)
        return Response(serializer.data)

    except Station.DoesNotExist:
        return Response(
            {'error': 'Estación no encontrada'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def latest_measurements_all_stations(request):
    """
    RF2.1: Obtiene las últimas mediciones de todas las estaciones

    Endpoint: GET /api/measurements/latest/
    """
    # Filtrar estaciones según permisos del usuario
    if request.user.role == UserRole.ADMIN:
        stations = Station.objects.filter(is_active=True)
    else:
        stations = request.user.assigned_stations.filter(is_active=True)

    latest_measurements = []

    for station in stations:
        latest = Measurement.objects.filter(
            station=station
        ).select_related('station', 'sensor').order_by('-timestamp').first()

        if latest:
            latest_measurements.append(latest)

    serializer = LatestMeasurementSerializer(latest_measurements, many=True)
    return Response(serializer.data)


# RF2.2 - Almacenamiento de Datos
class MeasurementCreateView(generics.CreateAPIView):
    """
    RF2.2: Endpoint principal para recibir mediciones desde sensores/PLC

    Endpoint: POST /api/measurements/
    """
    serializer_class = MeasurementCreateSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            measurement = serializer.save()
            response_serializer = MeasurementDetailSerializer(measurement)
            return Response(
                response_serializer.data,
                status=status.HTTP_201_CREATED
            )
        except ValueError as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def batch_create_measurements(request):
    """
    RF2.2: Endpoint para crear múltiples mediciones en lote (optimización para PLC)

    Endpoint: POST /api/measurements/batch/
    """
    serializer = BatchMeasurementCreateSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    try:
        result = serializer.save()
        return Response(
            {
                'message': f'Se crearon {result["count"]} mediciones exitosamente',
                'count': result['count']
            },
            status=status.HTTP_201_CREATED
        )
    except Exception as e:
        return Response(
            {'error': f'Error al procesar lote de mediciones: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )


# RF2.3 - Historial de Mediciones
class MeasurementListView(generics.ListAPIView):
    """
    RF2.3: Lista historial de mediciones con filtros

    Endpoint: GET /api/measurements/
    """
    serializer_class = MeasurementListSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = MeasurementPagination
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_class = MeasurementFilter
    ordering_fields = ['timestamp', 'value']
    ordering = ['-timestamp']

    def get_queryset(self):
        """Filtrar mediciones según permisos del usuario"""
        queryset = Measurement.objects.select_related('station', 'sensor')

        # Filtrar por permisos
        if self.request.user.role != UserRole.ADMIN:
            # Usuarios no admin solo ven mediciones de estaciones asignadas
            assigned_stations = self.request.user.assigned_stations.values_list('id', flat=True)
            queryset = queryset.filter(station__id__in=assigned_stations)

        return queryset


class MeasurementDetailView(generics.RetrieveAPIView):
    """
    Detalle de una medición específica

    Endpoint: GET /api/measurements/{id}/
    """
    serializer_class = MeasurementDetailSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filtrar mediciones según permisos del usuario"""
        queryset = Measurement.objects.select_related('station', 'sensor')

        if self.request.user.role != UserRole.ADMIN:
            assigned_stations = self.request.user.assigned_stations.values_list('id', flat=True)
            queryset = queryset.filter(station__id__in=assigned_stations)

        return queryset


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def measurement_statistics(request):
    """
    RF2.3: Estadísticas de mediciones por estación y tipo

    Endpoint: GET /api/measurements/statistics/
    """
    # Filtrar por permisos
    if request.user.role == UserRole.ADMIN:
        stations_filter = Q()
    else:
        assigned_stations = request.user.assigned_stations.values_list('id', flat=True)
        stations_filter = Q(station__id__in=assigned_stations)

    # Parámetros de filtro opcionales
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    station_id = request.GET.get('station_id')

    queryset = Measurement.objects.filter(stations_filter)

    if date_from:
        queryset = queryset.filter(timestamp__gte=date_from)
    if date_to:
        queryset = queryset.filter(timestamp__lte=date_to)
    if station_id:
        queryset = queryset.filter(station__id=station_id)

    # Agrupar y calcular estadísticas
    stats = queryset.values(
        'station__id', 'station__name', 'measurement_type', 'unit'
    ).annotate(
        count=Count('id'),
        avg_value=Avg('value'),
        min_value=Min('value'),
        max_value=Max('value'),
        latest_timestamp=Max('timestamp')
    ).order_by('station__name', 'measurement_type')

    # Añadir display names
    for stat in stats:
        stat['measurement_type_display'] = dict(MeasurementType.choices).get(
            stat['measurement_type'], stat['measurement_type']
        )

    serializer = MeasurementStatsSerializer(stats, many=True)
    return Response(serializer.data)


# RF2.4 - Configuración de Umbrales
class ThresholdListCreateView(generics.ListCreateAPIView):
    """
    RF2.4: Lista y crea umbrales de alerta

    Endpoint: GET/POST /api/thresholds/
    """
    serializer_class = ThresholdSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filtrar umbrales según permisos del usuario"""
        queryset = Threshold.objects.select_related('station', 'created_by', 'updated_by')

        if self.request.user.role != UserRole.ADMIN:
            assigned_stations = self.request.user.assigned_stations.values_list('id', flat=True)
            queryset = queryset.filter(station__id__in=assigned_stations)

        # Filtro opcional por estación
        station_id = self.request.query_params.get('station_id')
        if station_id:
            queryset = queryset.filter(station__id=station_id)

        return queryset.order_by('station__name', 'measurement_type')

    def perform_create(self, serializer):
        """Solo admins y técnicos pueden crear umbrales"""
        if self.request.user.role == UserRole.OBSERVER:
            raise PermissionError("No tiene permisos para crear umbrales")

        # Verificar que puede administrar la estación
        station = serializer.validated_data['station']
        if (self.request.user.role != UserRole.ADMIN and
                not station.assigned_users.filter(id=self.request.user.id).exists()):
            raise PermissionError("No tiene permisos para configurar umbrales en esta estación")

        serializer.save()


class ThresholdDetailView(generics.RetrieveUpdateDestroyAPIView):
    """
    RF2.4: Detalle, actualización y eliminación de umbrales

    Endpoint: GET/PUT/PATCH/DELETE /api/thresholds/{id}/
    """
    serializer_class = ThresholdSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filtrar umbrales según permisos del usuario"""
        queryset = Threshold.objects.select_related('station', 'created_by', 'updated_by')

        if self.request.user.role != UserRole.ADMIN:
            assigned_stations = self.request.user.assigned_stations.values_list('id', flat=True)
            queryset = queryset.filter(station__id__in=assigned_stations)

        return queryset

    def perform_update(self, serializer):
        """Solo admins y técnicos pueden modificar umbrales"""
        if self.request.user.role == UserRole.OBSERVER:
            raise PermissionError("No tiene permisos para modificar umbrales")

        serializer.save()

    def perform_destroy(self, instance):
        """Solo admins pueden eliminar umbrales"""
        if self.request.user.role != UserRole.ADMIN:
            raise PermissionError("Solo administradores pueden eliminar umbrales")

        instance.delete()


# RF2.5 - Alertas
class AlertListView(generics.ListAPIView):
    """
    RF2.5: Lista alertas del sistema

    Endpoint: GET /api/alerts/
    """
    serializer_class = AlertSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = MeasurementPagination
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    ordering_fields = ['triggered_at', 'level']
    ordering = ['-triggered_at']

    def get_queryset(self):
        """Filtrar alertas según permisos del usuario"""
        queryset = Alert.objects.select_related(
            'station', 'measurement', 'threshold',
            'acknowledged_by', 'resolved_by'
        )

        if self.request.user.role != UserRole.ADMIN:
            assigned_stations = self.request.user.assigned_stations.values_list('id', flat=True)
            queryset = queryset.filter(station__id__in=assigned_stations)

        # Filtros opcionales
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        level_filter = self.request.query_params.get('level')
        if level_filter:
            queryset = queryset.filter(level=level_filter)

        station_id = self.request.query_params.get('station_id')
        if station_id:
            queryset = queryset.filter(station__id=station_id)

        return queryset


class AlertDetailView(generics.RetrieveAPIView):
    """
    Detalle de una alerta específica

    Endpoint: GET /api/alerts/{id}/
    """
    serializer_class = AlertSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filtrar alertas según permisos del usuario"""
        queryset = Alert.objects.select_related(
            'station', 'measurement', 'threshold',
            'acknowledged_by', 'resolved_by'
        )

        if self.request.user.role != UserRole.ADMIN:
            assigned_stations = self.request.user.assigned_stations.values_list('id', flat=True)
            queryset = queryset.filter(station__id__in=assigned_stations)

        return queryset


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def alert_action(request, alert_id):
    """
    RF2.5: Acciones sobre alertas (reconocer, resolver, descartar)

    Endpoint: POST /api/alerts/{id}/action/
    """
    try:
        alert = get_object_or_404(Alert, id=alert_id)

        # Verificar permisos
        if (request.user.role != UserRole.ADMIN and
                not alert.station.assigned_users.filter(id=request.user.id).exists()):
            return Response(
                {'error': 'No tiene permisos para gestionar esta alerta'},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = AlertActionSerializer(data=request.data, context={'alert': alert})
        serializer.is_valid(raise_exception=True)

        action = serializer.validated_data['action']
        notes = serializer.validated_data.get('notes', '')

        # Ejecutar acción
        if action == 'acknowledge':
            alert.acknowledge(request.user, notes)
        elif action == 'resolve':
            alert.resolve(request.user, notes)
        elif action == 'dismiss':
            alert.dismiss(request.user, notes)

        # Retornar alerta actualizada
        response_serializer = AlertSerializer(alert)
        return Response(response_serializer.data)

    except Alert.DoesNotExist:
        return Response(
            {'error': 'Alerta no encontrada'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def active_alerts_summary(request):
    """
    RF2.5: Resumen de alertas activas por estación

    Endpoint: GET /api/alerts/active-summary/
    """
    # Filtrar por permisos
    if request.user.role == UserRole.ADMIN:
        stations_filter = Q()
    else:
        assigned_stations = request.user.assigned_stations.values_list('id', flat=True)
        stations_filter = Q(station__id__in=assigned_stations)

    # Contar alertas activas por nivel y estación
    active_alerts = Alert.objects.filter(
        stations_filter,
        status=AlertStatus.ACTIVE
    ).values(
        'station__id', 'station__name', 'level'
    ).annotate(
        count=Count('id')
    ).order_by('station__name', 'level')

    # Organizar datos por estación
    summary = {}
    for alert_data in active_alerts:
        station_id = alert_data['station__id']
        station_name = alert_data['station__name']

        if station_id not in summary:
            summary[station_id] = {
                'station_id': station_id,
                'station_name': station_name,
                'total_alerts': 0,
                'by_level': {
                    'info': 0,
                    'warning': 0,
                    'critical': 0,
                    'emergency': 0
                }
            }

        level = alert_data['level']
        count = alert_data['count']
        summary[station_id]['by_level'][level] = count
        summary[station_id]['total_alerts'] += count

    return Response(list(summary.values()))


# Configuración de Mediciones
class MeasurementConfigurationListCreateView(generics.ListCreateAPIView):
    """
    Lista y crea configuraciones de medición

    Endpoint: GET/POST /api/measurement-configurations/
    """
    serializer_class = MeasurementConfigurationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filtrar configuraciones según permisos del usuario"""
        queryset = MeasurementConfiguration.objects.select_related('station')

        if self.request.user.role != UserRole.ADMIN:
            assigned_stations = self.request.user.assigned_stations.values_list('id', flat=True)
            queryset = queryset.filter(station__id__in=assigned_stations)

        return queryset.order_by('station__name')

    def perform_create(self, serializer):
        """Solo admins pueden crear configuraciones"""
        if self.request.user.role != UserRole.ADMIN:
            raise PermissionError("Solo administradores pueden crear configuraciones")

        serializer.save()


class MeasurementConfigurationDetailView(generics.RetrieveUpdateAPIView):
    """
    Detalle y actualización de configuraciones de medición

    Endpoint: GET/PUT/PATCH /api/measurement-configurations/{id}/
    """
    serializer_class = MeasurementConfigurationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Filtrar configuraciones según permisos del usuario"""
        queryset = MeasurementConfiguration.objects.select_related('station')

        if self.request.user.role != UserRole.ADMIN:
            assigned_stations = self.request.user.assigned_stations.values_list('id', flat=True)
            queryset = queryset.filter(station__id__in=assigned_stations)

        return queryset

    def perform_update(self, serializer):
        """Solo admins pueden modificar configuraciones"""
        if self.request.user.role != UserRole.ADMIN:
            raise PermissionError("Solo administradores pueden modificar configuraciones")

        serializer.save()


# ===================================================================
# MÓDULO 3: REPORTES (RF3.1, RF3.2, RF3.3)
# ===================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def daily_average_report(request):
    """
    RF3.1: Reporte de Promedios Diarios

    Genera un reporte con los promedios diarios de mediciones
    para un rango de fechas y estaciones específicas.

    Endpoint: GET /api/measurements/reports/daily-averages/

    Parámetros:
    - date_from: Fecha desde (YYYY-MM-DD)
    - date_to: Fecha hasta (YYYY-MM-DD)
    - station_id: ID de estación (opcional, si no se especifica incluye todas)
    - measurement_type: Tipo de medición (water_level, flow_rate, temperature, ph)
    """
    # Validar parámetros requeridos
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    measurement_type = request.GET.get('measurement_type', 'water_level')

    if not date_from or not date_to:
        return Response(
            {'error': 'Los parámetros date_from y date_to son requeridos (formato: YYYY-MM-DD)'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # Convertir fechas
        date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()

        if date_from_obj > date_to_obj:
            return Response(
                {'error': 'La fecha_desde no puede ser mayor que fecha_hasta'},
                status=status.HTTP_400_BAD_REQUEST
            )

    except ValueError:
        return Response(
            {'error': 'Formato de fecha inválido. Use YYYY-MM-DD'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Filtrar por permisos del usuario
    if request.user.role == UserRole.ADMIN:
        stations_filter = Q()
    else:
        assigned_stations = request.user.assigned_stations.values_list('id', flat=True)
        stations_filter = Q(station__id__in=assigned_stations)

    # Construir consulta base
    queryset = Measurement.objects.filter(
        stations_filter,
        measurement_type=measurement_type,
        timestamp__date__gte=date_from_obj,
        timestamp__date__lte=date_to_obj,
        quality_flag='good'  # Solo mediciones de buena calidad
    )

    # Filtro opcional por estación
    station_id = request.GET.get('station_id')
    if station_id:
        try:
            station_id = int(station_id)
            queryset = queryset.filter(station__id=station_id)
        except ValueError:
            return Response(
                {'error': 'station_id debe ser un número entero'},
                status=status.HTTP_400_BAD_REQUEST
            )

    # Agregar por día y estación
    daily_averages = queryset.extra(
        select={'date': 'DATE(timestamp)'}
    ).values(
        'date', 'station__id', 'station__name', 'station__code', 'unit'
    ).annotate(
        avg_value=Avg('value'),
        min_value=Min('value'),
        max_value=Max('value'),
        count_measurements=Count('id'),
        first_measurement_time=Min('timestamp'),
        last_measurement_time=Max('timestamp')
    ).order_by('date', 'station__name')

    # Formatear datos para el serializer
    report_data = []
    for daily_avg in daily_averages:
        report_data.append({
            'date': daily_avg['date'],
            'station_id': daily_avg['station__id'],
            'station_name': daily_avg['station__name'],
            'station_code': daily_avg['station__code'],
            'measurement_type': measurement_type,
            'measurement_type_display': dict(MeasurementType.choices).get(measurement_type, measurement_type),
            'avg_value': daily_avg['avg_value'],
            'min_value': daily_avg['min_value'],
            'max_value': daily_avg['max_value'],
            'count': daily_avg['count_measurements'],
            'unit': daily_avg['unit'],
            'first_measurement_time': daily_avg['first_measurement_time'],
            'last_measurement_time': daily_avg['last_measurement_time']
        })

    serializer = DailyAverageReportSerializer(report_data, many=True)

    return Response({
        'report_info': {
            'type': 'daily_averages',
            'date_from': date_from,
            'date_to': date_to,
            'measurement_type': measurement_type,
            'station_filter': station_id,
            'total_records': len(report_data)
        },
        'results': serializer.data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def critical_events_report(request):
    """
    RF3.2: Reporte de Eventos Críticos

    Genera un reporte de todas las mediciones que han superado
    umbrales críticos o de advertencia en un período específico.

    Endpoint: GET /api/measurements/reports/critical-events/

    Parámetros:
    - date_from: Fecha desde (YYYY-MM-DD)
    - date_to: Fecha hasta (YYYY-MM-DD)
    - station_id: ID de estación (opcional)
    - level: Nivel de umbral (warning, critical, emergency - opcional)
    """
    # Validar parámetros requeridos
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if not date_from or not date_to:
        return Response(
            {'error': 'Los parámetros date_from y date_to son requeridos (formato: YYYY-MM-DD)'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # Convertir fechas
        date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()

        if date_from_obj > date_to_obj:
            return Response(
                {'error': 'La fecha_desde no puede ser mayor que fecha_hasta'},
                status=status.HTTP_400_BAD_REQUEST
            )

    except ValueError:
        return Response(
            {'error': 'Formato de fecha inválido. Use YYYY-MM-DD'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Filtrar por permisos del usuario
    if request.user.role == UserRole.ADMIN:
        stations_filter = Q()
    else:
        assigned_stations = request.user.assigned_stations.values_list('id', flat=True)
        stations_filter = Q(station__id__in=assigned_stations)

    # Construir consulta base para alertas generadas en el período
    alerts_queryset = Alert.objects.filter(
        stations_filter,
        triggered_at__date__gte=date_from_obj,
        triggered_at__date__lte=date_to_obj
    ).select_related(
        'station', 'measurement', 'threshold'
    )

    # Filtros opcionales
    station_id = request.GET.get('station_id')
    if station_id:
        try:
            station_id = int(station_id)
            alerts_queryset = alerts_queryset.filter(station__id=station_id)
        except ValueError:
            return Response(
                {'error': 'station_id debe ser un número entero'},
                status=status.HTTP_400_BAD_REQUEST
            )

    level_filter = request.GET.get('level')
    if level_filter and level_filter in ['warning', 'critical', 'emergency']:
        alerts_queryset = alerts_queryset.filter(level=level_filter)

    # Obtener alertas ordenadas por fecha
    alerts = alerts_queryset.order_by('-triggered_at')

    # Formatear datos para el serializer
    report_data = []
    for alert in alerts:
        measurement = alert.measurement
        threshold = alert.threshold

        report_data.append({
            'event_id': alert.id,
            'timestamp': alert.triggered_at,
            'station_id': alert.station.id,
            'station_name': alert.station.name,
            'station_code': alert.station.code,
            'measurement_type': measurement.measurement_type if measurement else threshold.measurement_type,
            'measurement_type_display': dict(MeasurementType.choices).get(
                measurement.measurement_type if measurement else threshold.measurement_type,
                measurement.measurement_type if measurement else threshold.measurement_type
            ),
            'measured_value': str(measurement.value) if measurement else None,
            'unit': measurement.unit if measurement else threshold.unit,
            'threshold_level': alert.level,
            'threshold_level_display': alert.get_level_display(),
            'threshold_exceeded': {
                'warning_min': str(threshold.warning_min) if threshold.warning_min else None,
                'warning_max': str(threshold.warning_max) if threshold.warning_max else None,
                'critical_min': str(threshold.critical_min) if threshold.critical_min else None,
                'critical_max': str(threshold.critical_max) if threshold.critical_max else None,
            },
            'alert_status': alert.status,
            'alert_status_display': alert.get_status_display(),
            'alert_title': alert.title,
            'alert_message': alert.message,
            'duration_minutes': alert.duration.total_seconds() / 60 if alert.status in ['resolved', 'dismissed'] else None
        })

    serializer = CriticalEventsReportSerializer(report_data, many=True)

    return Response({
        'report_info': {
            'type': 'critical_events',
            'date_from': date_from,
            'date_to': date_to,
            'station_filter': station_id,
            'level_filter': level_filter,
            'total_events': len(report_data)
        },
        'results': serializer.data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def comparative_report(request):
    """
    RF3.3: Reporte Comparativo entre Estaciones

    Genera un reporte comparativo de mediciones entre múltiples estaciones
    para un período específico, ideal para gráficos y análisis comparativo.

    Endpoint: GET /api/measurements/reports/comparative/

    Parámetros:
    - date_from: Fecha desde (YYYY-MM-DD)
    - date_to: Fecha hasta (YYYY-MM-DD)
    - stations: IDs de estaciones separados por coma (ej: 1,2,3)
    - measurement_type: Tipo de medición (water_level, flow_rate, temperature, ph)
    - aggregation: Tipo de agregación (daily, hourly - por defecto daily)
    """
    # Validar parámetros requeridos
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    stations_param = request.GET.get('stations')
    measurement_type = request.GET.get('measurement_type', 'water_level')
    aggregation = request.GET.get('aggregation', 'daily')

    if not date_from or not date_to or not stations_param:
        return Response(
            {'error': 'Los parámetros date_from, date_to y stations son requeridos'},
            status=status.HTTP_400_BAD_REQUEST
        )

    if aggregation not in ['daily', 'hourly']:
        return Response(
            {'error': 'aggregation debe ser "daily" o "hourly"'},
            status=status.HTTP_400_BAD_REQUEST
        )

    try:
        # Convertir fechas
        date_from_obj = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to_obj = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()

        if date_from_obj > date_to_obj:
            return Response(
                {'error': 'La fecha_desde no puede ser mayor que fecha_hasta'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Parsear IDs de estaciones
        station_ids = [int(id.strip()) for id in stations_param.split(',') if id.strip()]

        if len(station_ids) < 2:
            return Response(
                {'error': 'Se requieren al menos 2 estaciones para el reporte comparativo'},
                status=status.HTTP_400_BAD_REQUEST
            )

    except ValueError:
        return Response(
            {'error': 'Formato inválido en parámetros de fecha o IDs de estaciones'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Verificar permisos del usuario para las estaciones solicitadas
    if request.user.role == UserRole.ADMIN:
        accessible_stations = Station.objects.filter(id__in=station_ids, is_active=True)
    else:
        accessible_stations = request.user.assigned_stations.filter(
            id__in=station_ids, is_active=True
        )

    accessible_station_ids = list(accessible_stations.values_list('id', flat=True))

    if not accessible_station_ids:
        return Response(
            {'error': 'No tiene permisos para acceder a las estaciones solicitadas'},
            status=status.HTTP_403_FORBIDDEN
        )

    # Construir consulta base
    queryset = Measurement.objects.filter(
        station__id__in=accessible_station_ids,
        measurement_type=measurement_type,
        timestamp__date__gte=date_from_obj,
        timestamp__date__lte=date_to_obj,
        quality_flag='good'  # Solo mediciones de buena calidad
    )

    # Agregar según el tipo de agregación
    if aggregation == 'daily':
        date_trunc_sql = 'DATE(timestamp)'
        time_format = '%Y-%m-%d'
    else:  # hourly
        date_trunc_sql = 'DATE_FORMAT(timestamp, "%Y-%m-%d %H:00:00")'
        time_format = '%Y-%m-%d %H:00:00'

    # Realizar agregación
    comparative_data = queryset.extra(
        select={'time_period': date_trunc_sql}
    ).values(
        'time_period', 'station__id', 'station__name', 'station__code', 'unit'
    ).annotate(
        avg_value=Avg('value'),
        min_value=Min('value'),
        max_value=Max('value'),
        count_measurements=Count('id')
    ).order_by('time_period', 'station__name')

    # Formatear datos para el serializer ComparativeReportSerializer
    # El serializer espera una estructura específica, creamos un objeto único con todos los datos

    # Crear datos por estación
    stations_data = []
    global_stats = {
        'total_measurements': 0,
        'avg_value': 0,
        'min_value': None,
        'max_value': None
    }

    # Agrupar por estación
    station_groups = {}
    for data in comparative_data:
        station_id = data['station__id']
        if station_id not in station_groups:
            station_groups[station_id] = {
                'station_id': station_id,
                'station_name': data['station__name'],
                'station_code': data['station__code'],
                'unit': data['unit'],
                'data_points': [],
                'statistics': {
                    'count': 0,
                    'avg': 0,
                    'min': None,
                    'max': None
                }
            }

        # Agregar estadísticas de esta estación
        station_stats = station_groups[station_id]['statistics']
        station_stats['count'] += data['count_measurements']
        if station_stats['avg'] == 0:
            station_stats['avg'] = float(data['avg_value']) if data['avg_value'] else 0
        if station_stats['min'] is None or (data['min_value'] and float(data['min_value']) < station_stats['min']):
            station_stats['min'] = float(data['min_value']) if data['min_value'] else 0
        if station_stats['max'] is None or (data['max_value'] and float(data['max_value']) > station_stats['max']):
            station_stats['max'] = float(data['max_value']) if data['max_value'] else 0

        # Actualizar estadísticas globales
        global_stats['total_measurements'] += data['count_measurements']
        if global_stats['min_value'] is None or (data['min_value'] and float(data['min_value']) < global_stats['min_value']):
            global_stats['min_value'] = float(data['min_value']) if data['min_value'] else 0
        if global_stats['max_value'] is None or (data['max_value'] and float(data['max_value']) > global_stats['max_value']):
            global_stats['max_value'] = float(data['max_value']) if data['max_value'] else 0

    # Convertir grupos a lista
    for station_data in station_groups.values():
        stations_data.append(station_data)

    # Calcular promedio global
    if stations_data:
        total_avg = sum(station['statistics']['avg'] for station in stations_data)
        global_stats['avg_value'] = total_avg / len(stations_data)

    # Crear datos para el serializer
    report_data = {
        'measurement_type': measurement_type,
        'measurement_type_display': dict(MeasurementType.choices).get(measurement_type, measurement_type),
        'period_start': f"{date_from_obj}T00:00:00Z",
        'period_end': f"{date_to_obj}T23:59:59Z",
        'total_stations': len(accessible_stations),
        'stations_data': stations_data,
        'global_statistics': global_stats
    }

    serializer = ComparativeReportSerializer(report_data)

    return Response({
        'report_info': {
            'type': 'comparative',
            'date_from': date_from,
            'date_to': date_to,
            'measurement_type': measurement_type,
            'aggregation': aggregation,
            'stations_included': [{'id': s.id, 'name': s.name, 'code': s.code, 'location': s.location} for s in accessible_stations],
            'total_records': len(comparative_data)
        },
        'results': serializer.data
    })


# ===================================================================
# EXPORT REPORTS (PDF/Excel)
# ===================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_report_pdf(request):
    """
    Exporta reportes a PDF

    Endpoint: GET /api/measurements/reports/export/pdf/

    Parámetros:
    - report_type: Tipo de reporte (daily-averages, critical-events, comparative)
    - date_from: Fecha desde (YYYY-MM-DD)
    - date_to: Fecha hasta (YYYY-MM-DD)
    - station_id: ID de estación (opcional)
    - measurement_type: Tipo de medición (opcional)
    - level: Nivel de alerta (para critical-events, opcional)
    - stations: IDs separados por coma (para comparative)
    - aggregation: daily/hourly (para comparative, opcional)
    """
    report_type = request.GET.get('report_type')
    if not report_type:
        return Response(
            {'error': 'El parámetro report_type es requerido'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Obtener datos del reporte según el tipo
    if report_type == 'daily-averages':
        response = daily_average_report(request._request)
        if response.status_code != 200:
            return response
        data = response.data
        report_data = data['results']
        report_info = data['report_info']

    elif report_type == 'critical-events':
        response = critical_events_report(request._request)
        if response.status_code != 200:
            return response
        data = response.data
        report_data = data['results']
        report_info = data['report_info']

    elif report_type == 'comparative':
        response = comparative_report(request._request)
        if response.status_code != 200:
            return response
        data = response.data
        report_data = data['results']
        report_info = data['report_info']

    else:
        return Response(
            {'error': 'Tipo de reporte no válido. Use: daily-averages, critical-events, comparative'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Crear buffer para el PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Título del reporte
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center
    )

    report_titles = {
        'daily-averages': 'Reporte de Promedios Diarios',
        'critical-events': 'Reporte de Eventos Críticos',
        'comparative': 'Reporte Comparativo entre Estaciones'
    }

    title = Paragraph(report_titles.get(report_type, 'Reporte'), title_style)
    elements.append(title)

    # Información del reporte
    info_style = styles['Normal']
    info_text = f"""
    <b>Período:</b> {report_info.get('date_from', 'N/A')} - {report_info.get('date_to', 'N/A')}<br/>
    <b>Tipo de Medición:</b> {report_info.get('measurement_type', 'N/A')}<br/>
    <b>Total de Registros:</b> {report_info.get('total_records', 0)}<br/>
    """

    if report_type == 'critical-events' and report_info.get('level_filter'):
        info_text += f"<b>Nivel de Alerta:</b> {report_info['level_filter']}<br/>"

    if report_info.get('station_filter'):
        info_text += f"<b>Estación:</b> {report_info['station_filter']}<br/>"

    elements.append(Paragraph(info_text, info_style))
    elements.append(Spacer(1, 20))

    # Crear tabla según el tipo de reporte
    if report_type == 'daily-averages':
        headers = ['Fecha', 'Estación', 'Código', 'Promedio', 'Mínimo', 'Máximo', 'Conteo', 'Unidad']
        table_data = [headers]

        for item in report_data:
            row = [
                str(item['date']),
                item['station_name'],
                item['station_code'],
                f"{item['avg_value']:.2f}" if item['avg_value'] else 'N/A',
                f"{item['min_value']:.2f}" if item['min_value'] else 'N/A',
                f"{item['max_value']:.2f}" if item['max_value'] else 'N/A',
                str(item['count']),
                item['unit']
            ]
            table_data.append(row)

    elif report_type == 'critical-events':
        headers = ['Fecha/Hora', 'Estación', 'Tipo', 'Valor', 'Unidad', 'Nivel', 'Estado', 'Mensaje']
        table_data = [headers]

        for item in report_data:
            row = [
                item['timestamp'].strftime('%Y-%m-%d %H:%M') if hasattr(item['timestamp'], 'strftime') else str(item['timestamp']),
                item['station_name'],
                item['measurement_type_display'],
                item['measured_value'] or 'N/A',
                item['unit'],
                item['threshold_level_display'],
                item['alert_status_display'],
                item['alert_message'][:50] + '...' if len(item['alert_message']) > 50 else item['alert_message']
            ]
            table_data.append(row)

    elif report_type == 'comparative':
        # Para reporte comparativo, mostrar estadísticas por estación
        headers = ['Estación', 'Código', 'Conteo Total', 'Promedio', 'Mínimo', 'Máximo', 'Unidad']
        table_data = [headers]

        for station in report_data['stations_data']:
            stats = station['statistics']
            row = [
                station['station_name'],
                station['station_code'],
                str(stats['count']),
                f"{stats['avg']:.2f}" if stats['avg'] else 'N/A',
                f"{stats['min']:.2f}" if stats['min'] is not None else 'N/A',
                f"{stats['max']:.2f}" if stats['max'] is not None else 'N/A',
                station['unit']
            ]
            table_data.append(row)

    # Crear tabla
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)

    # Generar PDF
    doc.build(elements)

    # Preparar respuesta
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{datetime.date.today()}.pdf"'

    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_report_excel(request):
    """
    Exporta reportes a Excel

    Endpoint: GET /api/measurements/reports/export/excel/

    Parámetros: Mismos que export_report_pdf
    """
    report_type = request.GET.get('report_type')
    if not report_type:
        return Response(
            {'error': 'El parámetro report_type es requerido'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Obtener datos del reporte según el tipo (mismo código que PDF)
    if report_type == 'daily-averages':
        response = daily_average_report(request._request)
        if response.status_code != 200:
            return response
        data = response.data
        report_data = data['results']
        report_info = data['report_info']

    elif report_type == 'critical-events':
        response = critical_events_report(request._request)
        if response.status_code != 200:
            return response
        data = response.data
        report_data = data['results']
        report_info = data['report_info']

    elif report_type == 'comparative':
        response = comparative_report(request._request)
        if response.status_code != 200:
            return response
        data = response.data
        report_data = data['results']
        report_info = data['report_info']

    else:
        return Response(
            {'error': 'Tipo de reporte no válido. Use: daily-averages, critical-events, comparative'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Crear workbook de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.replace('-', '_').title()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Título del reporte
    report_titles = {
        'daily-averages': 'Reporte de Promedios Diarios',
        'critical-events': 'Reporte de Eventos Críticos',
        'comparative': 'Reporte Comparativo entre Estaciones'
    }

    ws['A1'] = report_titles.get(report_type, 'Reporte')
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')

    # Información del reporte
    row = 3
    ws[f'A{row}'] = f'Período: {report_info.get("date_from", "N/A")} - {report_info.get("date_to", "N/A")}'
    row += 1
    ws[f'A{row}'] = f'Tipo de Medición: {report_info.get("measurement_type", "N/A")}'
    row += 1
    ws[f'A{row}'] = f'Total de Registros: {report_info.get("total_records", 0)}'
    row += 1

    if report_type == 'critical-events' and report_info.get('level_filter'):
        ws[f'A{row}'] = f'Nivel de Alerta: {report_info["level_filter"]}'
        row += 1

    if report_info.get('station_filter'):
        ws[f'A{row}'] = f'Estación: {report_info["station_filter"]}'
        row += 1

    # Espacio
    row += 2

    # Headers y datos según tipo de reporte
    if report_type == 'daily-averages':
        headers = ['Fecha', 'Estación', 'Código', 'Promedio', 'Mínimo', 'Máximo', 'Conteo', 'Unidad']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row += 1

        for item in report_data:
            ws.cell(row=row, column=1, value=str(item['date']))
            ws.cell(row=row, column=2, value=item['station_name'])
            ws.cell(row=row, column=3, value=item['station_code'])
            ws.cell(row=row, column=4, value=round(item['avg_value'], 2) if item['avg_value'] else None)
            ws.cell(row=row, column=5, value=round(item['min_value'], 2) if item['min_value'] else None)
            ws.cell(row=row, column=6, value=round(item['max_value'], 2) if item['max_value'] else None)
            ws.cell(row=row, column=7, value=item['count'])
            ws.cell(row=row, column=8, value=item['unit'])
            row += 1

    elif report_type == 'critical-events':
        headers = ['Fecha/Hora', 'Estación', 'Tipo', 'Valor', 'Unidad', 'Nivel', 'Estado', 'Mensaje']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row += 1

        for item in report_data:
            timestamp = item['timestamp']
            if hasattr(timestamp, 'strftime'):
                timestamp_str = timestamp.strftime('%Y-%m-%d %H:%M')
            else:
                timestamp_str = str(timestamp)

            ws.cell(row=row, column=1, value=timestamp_str)
            ws.cell(row=row, column=2, value=item['station_name'])
            ws.cell(row=row, column=3, value=item['measurement_type_display'])
            ws.cell(row=row, column=4, value=item['measured_value'] or 'N/A')
            ws.cell(row=row, column=5, value=item['unit'])
            ws.cell(row=row, column=6, value=item['threshold_level_display'])
            ws.cell(row=row, column=7, value=item['alert_status_display'])
            ws.cell(row=row, column=8, value=item['alert_message'])
            row += 1

    elif report_type == 'comparative':
        headers = ['Estación', 'Código', 'Conteo Total', 'Promedio', 'Mínimo', 'Máximo', 'Unidad']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        row += 1

        for station in report_data['stations_data']:
            stats = station['statistics']
            ws.cell(row=row, column=1, value=station['station_name'])
            ws.cell(row=row, column=2, value=station['station_code'])
            ws.cell(row=row, column=3, value=stats['count'])
            ws.cell(row=row, column=4, value=round(stats['avg'], 2) if stats['avg'] else None)
            ws.cell(row=row, column=5, value=round(stats['min'], 2) if stats['min'] is not None else None)
            ws.cell(row=row, column=6, value=round(stats['max'], 2) if stats['max'] is not None else None)
            ws.cell(row=row, column=7, value=station['unit'])
            row += 1

    # Ajustar ancho de columnas
    for col in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    # Crear respuesta
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{datetime.date.today()}.xlsx"'

    return response
